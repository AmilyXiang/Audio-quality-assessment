#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成问题文件的Excel报告
将NOK文件的问题帧信息输出到Excel表格中
"""

import json
import argparse
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def merge_intervals(intervals):
    """
    合并重叠的时间区间
    
    Args:
        intervals: [(start, end), ...] 时间区间列表
    
    Returns:
        list: 合并后的区间列表
    """
    if not intervals:
        return []
    
    # 按起始时间排序
    intervals = sorted(intervals, key=lambda x: x[0])
    merged = [intervals[0]]
    
    for current in intervals[1:]:
        last = merged[-1]
        # 如果当前区间与上一个区间重叠或相邻，合并
        if current[0] <= last[1]:
            merged[-1] = (last[0], max(last[1], current[1]))
        else:
            merged.append(current)
    
    return merged


def subtract_intervals(problem_interval, normal_intervals):
    """
    计算问题区间减去正常区间的差集
    
    Args:
        problem_interval: (start, end) 问题帧的时间窗口
        normal_intervals: [(start, end), ...] 正常帧的时间窗口列表
    
    Returns:
        list: 差集区间列表 [(start, end), ...]
    """
    if not normal_intervals:
        return [problem_interval]
    
    # 合并所有正常区间
    normal_intervals = merge_intervals(normal_intervals)
    
    # 从问题区间中逐个减去正常区间
    result = [problem_interval]
    
    for normal_start, normal_end in normal_intervals:
        new_result = []
        
        for prob_start, prob_end in result:
            # 情况1: 正常区间完全在问题区间之外
            if normal_end <= prob_start or normal_start >= prob_end:
                new_result.append((prob_start, prob_end))
            # 情况2: 正常区间完全覆盖问题区间
            elif normal_start <= prob_start and normal_end >= prob_end:
                pass  # 问题区间被完全覆盖，不添加
            # 情况3: 正常区间覆盖问题区间的左侧
            elif normal_start <= prob_start and normal_end < prob_end:
                new_result.append((normal_end, prob_end))
            # 情况4: 正常区间覆盖问题区间的右侧
            elif normal_start > prob_start and normal_end >= prob_end:
                new_result.append((prob_start, normal_start))
            # 情况5: 正常区间在问题区间中间，分割成两段
            else:  # normal_start > prob_start and normal_end < prob_end
                new_result.append((prob_start, normal_start))
                new_result.append((normal_end, prob_end))
        
        result = new_result
        if not result:
            break
    
    return result


def extract_problem_frames(comparison_data, frame_threshold=-0.3):
    """
    使用差集分析提取真正的问题时间段
    
    Args:
        comparison_data: 对比JSON数据
        frame_threshold: 问题帧阈值（默认-0.3）
    
    Returns:
        dict: {维度名: {'frames': [frame_ids], 'time_ranges': [时间段字符串]}}
    """
    filename = comparison_data.get('test_file', 'Unknown')
    status = comparison_data.get('status', 'Unknown')
    
    if status != 'NOK':
        return None
    
    problem_data = {
        'filename': filename,
        'MOS': {'frames': [], 'time_ranges': []},
        'NOI': {'frames': [], 'time_ranges': []},
        'DIS': {'frames': [], 'time_ranges': []},
        'COL': {'frames': [], 'time_ranges': []},
        'LOUD': {'frames': [], 'time_ranges': []}
    }
    
    # 遍历各个维度的帧差异
    metrics = comparison_data.get('metrics', {})
    
    for dim in ['mos', 'noi', 'dis', 'col', 'loud']:
        dim_key = dim.upper()
        dim_data = metrics.get(dim, {})
        frame_diffs = dim_data.get('frame_diffs', [])
        
        # 分离问题帧和正常帧
        problem_frames = []
        normal_frames = []
        
        for frame_info in frame_diffs:
            diff = frame_info.get('diff', 0)
            frame_id = frame_info.get('frame_id', 0)
            time_start = frame_info.get('time_start', 0)
            time_end = frame_info.get('time_end', time_start + 15.0)
            
            frame_data = {
                'frame_id': frame_id,
                'time_start': time_start,
                'time_end': time_end,
                'diff': diff
            }
            
            if diff < frame_threshold:
                problem_frames.append(frame_data)
            else:
                normal_frames.append(frame_data)
        
        if problem_frames:
            # 提取frame IDs
            problem_data[dim_key]['frames'] = [f['frame_id'] for f in problem_frames]
            
            # 使用差集分析计算真正的问题时间段
            problem_regions = compute_problem_regions(problem_frames, normal_frames)
            problem_data[dim_key]['time_ranges'] = problem_regions
    
    # 如果没有任何问题帧，返回None
    has_problems = any(len(v['frames']) > 0 for v in problem_data.values() if isinstance(v, dict))
    if not has_problems:
        return None
    
    return problem_data


def compute_problem_regions(problem_frames, normal_frames):
    """
    混合边界分析策略：
    - 孤立问题帧：与相邻OK帧比较边界差异
    - 连续问题帧：分析首尾差异
    
    Args:
        problem_frames: 问题帧列表
        normal_frames: 正常帧列表
    
    Returns:
        list: 问题时间段字符串列表 ['7.0-8.5s', '22.0-23.5s', ...]
    """
    if not problem_frames:
        return []
    
    # 构建所有帧的映射（按frame_id）
    all_frames = {f['frame_id']: f for f in problem_frames + normal_frames}
    
    # 按frame_id排序
    problem_frames = sorted(problem_frames, key=lambda x: x['frame_id'])
    
    # 识别连续问题帧组
    consecutive_groups = []
    current_group = [problem_frames[0]]
    
    for i in range(1, len(problem_frames)):
        if problem_frames[i]['frame_id'] == problem_frames[i-1]['frame_id'] + 1:
            current_group.append(problem_frames[i])
        else:
            consecutive_groups.append(current_group)
            current_group = [problem_frames[i]]
    consecutive_groups.append(current_group)
    
    # 对每组应用不同策略
    all_problem_regions = []
    
    for group in consecutive_groups:
        first_frame = group[0]   # 第一个问题帧
        last_frame = group[-1]   # 最后一个问题帧
        
        if len(group) == 1:
            # 孤立问题帧：与相邻OK帧比较边界
            prob_frame = first_frame
            prob_id = prob_frame['frame_id']
            boundary_regions = []
            
            # 检查前一帧（frame_id - 1）
            prev_id = prob_id - 1
            if prev_id in all_frames:
                prev_frame = all_frames[prev_id]
                if prev_frame['diff'] >= -0.3:  # 前一帧是OK的
                    # 问题帧相对于前一帧的新增后部
                    if prob_frame['time_end'] > prev_frame['time_end']:
                        boundary_regions.append((prev_frame['time_end'], prob_frame['time_end']))
            
            # 检查后一帧（frame_id + 1）
            next_id = prob_id + 1
            if next_id in all_frames:
                next_frame = all_frames[next_id]
                if next_frame['diff'] >= -0.3:  # 后一帧是OK的
                    # 问题帧相对于后一帧的新增前部
                    if prob_frame['time_start'] < next_frame['time_start']:
                        boundary_regions.append((prob_frame['time_start'], next_frame['time_start']))
            
            # 如果没有相邻OK帧，保留整个问题帧
            if not boundary_regions:
                boundary_regions.append((prob_frame['time_start'], prob_frame['time_end']))
            
            all_problem_regions.extend(boundary_regions)
        else:
            # 连续问题帧：分析首尾差异
            # 前部差异：第一个问题帧独有的前部（最后一个没覆盖）
            if first_frame['time_start'] < last_frame['time_start']:
                all_problem_regions.append((first_frame['time_start'], last_frame['time_start']))
            
            # 后部差异：最后一个问题帧独有的后部（第一个没覆盖）
            if last_frame['time_end'] > first_frame['time_end']:
                all_problem_regions.append((first_frame['time_end'], last_frame['time_end']))
    
    # 合并重叠区域
    if all_problem_regions:
        all_problem_regions = merge_intervals(all_problem_regions)
    
    # 格式化输出
    result = []
    for start, end in all_problem_regions:
        # 过滤掉过短的区域（小于0.5秒）
        if end - start >= 0.5:
            result.append(f"{start:.1f}-{end:.1f}s")
    
    return result


def append_to_excel(problem_data, excel_path, frame_threshold=-0.3):
    """
    将问题数据追加到Excel文件
    
    Args:
        problem_data: 问题帧数据
        excel_path: Excel文件路径
        frame_threshold: 阈值（用于标题）
    """
    if problem_data is None:
        return
    
    # 如果文件存在则加载，否则创建新文件
    if Path(excel_path).exists():
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "问题文件记录"
        
        # 创建表头
        headers = ['文件名', 'MOS', 'NOI', 'DIS', 'COL', 'LOUD']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 添加说明行
        ws.cell(row=2, column=1, value=f"阈值: {frame_threshold}")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        ws.cell(row=2, column=1).font = Font(italic=True, size=10)
        ws.cell(row=2, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # 找到下一个空行（跳过表头和说明行）
    next_row = ws.max_row + 1
    if next_row <= 2:
        next_row = 3
    
    # 提取文件名（去除路径和framewise_前缀）
    filename = problem_data['filename']
    if '/' in filename or '\\' in filename:
        filename = Path(filename).stem
    if filename.startswith('framewise_'):
        filename = filename[10:]
    
    # 写入Frame ID行
    ws.cell(row=next_row, column=1, value=filename)
    ws.cell(row=next_row, column=2, value=', '.join(map(str, problem_data['MOS']['frames'])) if problem_data['MOS']['frames'] else '-')
    ws.cell(row=next_row, column=3, value=', '.join(map(str, problem_data['NOI']['frames'])) if problem_data['NOI']['frames'] else '-')
    ws.cell(row=next_row, column=4, value=', '.join(map(str, problem_data['DIS']['frames'])) if problem_data['DIS']['frames'] else '-')
    ws.cell(row=next_row, column=5, value=', '.join(map(str, problem_data['COL']['frames'])) if problem_data['COL']['frames'] else '-')
    ws.cell(row=next_row, column=6, value=', '.join(map(str, problem_data['LOUD']['frames'])) if problem_data['LOUD']['frames'] else '-')
    
    # 设置Frame ID行样式
    for col in range(1, 7):
        cell = ws.cell(row=next_row, column=col)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        if col == 1:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # 写入时间行
    next_row += 1
    ws.cell(row=next_row, column=1, value=f"  ↳ 时间段")
    ws.cell(row=next_row, column=2, value=', '.join(problem_data['MOS']['time_ranges']) if problem_data['MOS']['time_ranges'] else '-')
    ws.cell(row=next_row, column=3, value=', '.join(problem_data['NOI']['time_ranges']) if problem_data['NOI']['time_ranges'] else '-')
    ws.cell(row=next_row, column=4, value=', '.join(problem_data['DIS']['time_ranges']) if problem_data['DIS']['time_ranges'] else '-')
    ws.cell(row=next_row, column=5, value=', '.join(problem_data['COL']['time_ranges']) if problem_data['COL']['time_ranges'] else '-')
    ws.cell(row=next_row, column=6, value=', '.join(problem_data['LOUD']['time_ranges']) if problem_data['LOUD']['time_ranges'] else '-')
    
    # 设置时间行样式
    for col in range(1, 7):
        cell = ws.cell(row=next_row, column=col)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.font = Font(italic=True, size=10, color="666666")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        if col == 1:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # 调整列宽
    ws.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 30
    
    # 保存文件
    wb.save(excel_path)


def process_single_file(json_path, excel_path, frame_threshold=-0.3):
    """
    处理单个JSON文件并追加到Excel
    
    Args:
        json_path: baseline_compare JSON文件路径
        excel_path: 输出Excel文件路径
        frame_threshold: 问题帧阈值
    
    Returns:
        bool: True表示添加了NOK记录，False表示文件为OK
    """
    # 读取JSON文件
    with open(json_path, 'r', encoding='utf-8') as f:
        comparison_data = json.load(f)
    
    # 检查状态
    status = comparison_data.get('status', 'Unknown')
    if status == 'OK':
        print(f"[跳过] {Path(json_path).name} - 文件质量OK")
        return False
    
    # 提取问题帧
    problem_data = extract_problem_frames(comparison_data, frame_threshold)
    
    if problem_data is None:
        print(f"[跳过] {Path(json_path).name} - 无问题帧")
        return False
    
    # 追加到Excel
    append_to_excel(problem_data, excel_path, frame_threshold)
    print(f"[已添加] {problem_data['filename']} -> {excel_path}")
    
    return True


def main():
    parser = argparse.ArgumentParser(description='生成问题文件Excel报告')
    parser.add_argument('input',
                       help='baseline_compare JSON文件或包含JSON文件的目录')
    parser.add_argument('--output', '-o', default='problem_files_report.xlsx',
                       help='输出Excel文件路径（默认: problem_files_report.xlsx）')
    parser.add_argument('--frame-threshold', type=float, default=-0.3,
                       help='问题帧差异阈值（默认-0.3）')
    
    args = parser.parse_args()
    
    input_path = Path(args.input)
    excel_path = Path(args.output)
    
    # 收集所有JSON文件
    json_files = []
    if input_path.is_file():
        if input_path.suffix == '.json' and 'baseline_compare_' in input_path.name:
            json_files.append(input_path)
        else:
            print(f"[错误] {input_path} 不是有效的baseline_compare JSON文件")
            return
    elif input_path.is_dir():
        json_files = list(input_path.glob('baseline_compare_*.json'))
    else:
        print(f"[错误] 路径不存在: {input_path}")
        return
    
    if not json_files:
        print("[错误] 未找到baseline_compare JSON文件")
        return
    
    print(f"[信息] 找到 {len(json_files)} 个对比文件")
    print(f"[信息] 问题帧阈值: {args.frame_threshold}")
    print(f"[信息] 输出文件: {excel_path}")
    print("")
    
    # 处理所有文件
    nok_count = 0
    ok_count = 0
    
    for json_file in sorted(json_files):
        if process_single_file(json_file, excel_path, args.frame_threshold):
            nok_count += 1
        else:
            ok_count += 1
    
    print("")
    print("=" * 80)
    print(f"[完成] 总计: {len(json_files)} 个文件")
    print(f"  - OK文件: {ok_count} 个（已跳过）")
    print(f"  - NOK文件: {nok_count} 个（已记录到Excel）")
    print(f"[保存] {excel_path}")
    print("=" * 80)


if __name__ == '__main__':
    main()
